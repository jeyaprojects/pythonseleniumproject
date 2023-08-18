import openpyxl


file="C:\excel\detail.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
rows=sheet.max_row
col=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value,end="  ")
    print()