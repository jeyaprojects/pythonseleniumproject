import openpyxl
from openpyxl.styles import PatternFill
'''
here we have all utilty function for all action

'''

def getrowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_row)

def getcolumncount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_column)

def readdata(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,colnum).value


def writedata(file,sheetname,rownum,colnum,datas):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,colnum).value= datas
    workbook.save(file)

def fillgreencolor(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greenFill=PatternFill(start_color='0000FF00',
                          end_color='0000FF00',
                          fill_type='solid')
    sheet.cell(rownum,colnum).fill=greenFill
    workbook.save(file)

def fillredcolor(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]

    redFill=PatternFill(start_color='00FF0000',
                          end_color='00FF0000',
                          fill_type='solid')
    sheet.cell(rownum,colnum).fill=redFill
    workbook.save(file)










